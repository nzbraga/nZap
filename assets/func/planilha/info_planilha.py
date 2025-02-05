from openpyxl import load_workbook
from assets.func.uteis.popUp import popUp

def info_planilha(index_pagina, nome_arquivo):

    if index_pagina == -1:
        popUp("Nenhuma pagina selecionada")
        return False
    planilha = load_workbook(f'{nome_arquivo}.xlsx')
    nome_pagina = planilha.sheetnames[index_pagina]
    pagina = planilha[nome_pagina]

    return pagina
